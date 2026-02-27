"""
Export router - Excel data export endpoints.
"""

from datetime import date
from typing import Optional, List
from uuid import UUID
from io import BytesIO

from fastapi import APIRouter, Depends, Query, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import cast, String
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.session import get_db
from app.core.auth import get_current_firebase_uid
from app.features.users.service import UserProfileService
from app.features.tenants.model import Tenant
from app.features.transactions.model import Transaction
from app.features.kosts.model import Kost

router = APIRouter()


from app.features.common.dependencies import get_current_user_region


def style_header_row(ws, row_num: int = 1):
    """Apply header styling to a row."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0f766d", end_color="0f766d", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row_num]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def auto_size_columns(ws):
    """Auto-size columns based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


@router.get("/excel")
async def export_to_excel(
    start_date: date = Query(..., description="Start date for export"),
    end_date: date = Query(..., description="End date for export"),
    data_types: List[str] = Query(..., description="Data types to export: tenants, payments, expenses"),
    region_id: Optional[UUID] = Depends(get_current_user_region),
    db: Session = Depends(get_db),
):
    """
    Export selected data types to a single Excel file with multiple sheets.
    Each data type becomes a separate sheet in the workbook.
    """
    
    if not data_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="At least one data type must be selected"
        )
    
    # Get all kost IDs in this region
    kost_ids = []
    if region_id:
        kosts = db.query(Kost).filter(Kost.region_id == region_id).all()
        kost_ids = [k.id for k in kosts]
    
    # Create workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Process each data type
    for data_type in data_types:
        if data_type == "tenants":
            _add_tenants_sheet(wb, db, kost_ids)
        elif data_type == "payments":
            _add_payments_sheet(wb, db, kost_ids, start_date, end_date)
        elif data_type == "expenses":
            _add_expenses_sheet(wb, db, kost_ids, start_date, end_date)
        elif data_type == "activity":
            # Activity log is no longer supported.
            continue
    
    # If no sheets were added (invalid data types), return error
    if len(wb.worksheets) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No valid data types selected"
        )
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"ekspor_data_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _add_tenants_sheet(wb: Workbook, db: Session, kost_ids: list):
    """Add tenants data sheet."""
    ws = wb.create_sheet("Data Penyewa")
    
    # Headers
    headers = ["Nama", "Telepon", "Nama Kost", "Tanggal Masuk", "Tanggal Keluar", "Harga Sewa", "Status"]
    ws.append(headers)
    style_header_row(ws)
    
    # Query tenants (all tenants, including inactive for full recap)
    query = db.query(Tenant)
    if kost_ids:
        query = query.filter(Tenant.kost_id.in_(kost_ids))
    
    tenants = query.order_by(Tenant.created_at.desc()).all()
    
    for tenant in tenants:
        # Get kost name
        kost = db.query(Kost).filter(Kost.id == tenant.kost_id).first()
        kost_name = kost.name if kost else "-"
        
        ws.append([
            tenant.name,
            tenant.phone or "-",
            kost_name,
            tenant.start_date.strftime("%d/%m/%Y") if tenant.start_date else "-",
            tenant.end_date.strftime("%d/%m/%Y") if tenant.end_date else "-",
            float(tenant.rent_price) if tenant.rent_price else 0,
            tenant.status or "-",
        ])
    
    auto_size_columns(ws)


def _add_payments_sheet(wb: Workbook, db: Session, kost_ids: list, start_date: date, end_date: date):
    """Add payments (income) data sheet."""
    ws = wb.create_sheet("Riwayat Pembayaran")
    
    # Headers
    headers = ["Tanggal", "Nama Penyewa", "Nama Kost", "Kategori", "Jumlah", "Keterangan"]
    ws.append(headers)
    style_header_row(ws)
    
    # Query income transactions
    query = db.query(Transaction).filter(
        cast(Transaction.type, String) == "income",
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date,
    )
    if kost_ids:
        query = query.filter(Transaction.kost_id.in_(kost_ids))
    
    transactions = query.order_by(Transaction.transaction_date.desc()).all()
    
    for tx in transactions:
        # Get names
        tenant_name = "-"
        if tx.tenant_id:
            tenant = db.query(Tenant).filter(Tenant.id == tx.tenant_id).first()
            tenant_name = tenant.name if tenant else "-"
        
        kost = db.query(Kost).filter(Kost.id == tx.kost_id).first()
        kost_name = kost.name if kost else "-"
        
        ws.append([
            tx.transaction_date.strftime("%d/%m/%Y"),
            tenant_name,
            kost_name,
            tx.category or "-",
            float(tx.amount),
            tx.description or "-",
        ])
    
    auto_size_columns(ws)


def _add_expenses_sheet(wb: Workbook, db: Session, kost_ids: list, start_date: date, end_date: date):
    """Add expenses data sheet."""
    ws = wb.create_sheet("Laporan Pengeluaran")
    
    # Headers
    headers = ["Tanggal", "Nama Kost", "Kategori", "Jumlah", "Keterangan"]
    ws.append(headers)
    style_header_row(ws)
    
    # Query expense transactions
    query = db.query(Transaction).filter(
        cast(Transaction.type, String) == "expense",
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date,
    )
    if kost_ids:
        query = query.filter(Transaction.kost_id.in_(kost_ids))
    
    transactions = query.order_by(Transaction.transaction_date.desc()).all()
    
    for tx in transactions:
        kost = db.query(Kost).filter(Kost.id == tx.kost_id).first()
        kost_name = kost.name if kost else "-"
        
        ws.append([
            tx.transaction_date.strftime("%d/%m/%Y"),
            kost_name,
            tx.category or "-",
            float(tx.amount),
            tx.description or "-",
        ])
    
    auto_size_columns(ws)


